# -*- coding: utf-8 -*-
"""
资源获取脚本：读 deepresearch 汇总表 → 直链 PDF 自动下载 → 每行写入 registry.csv。

用法：
  python fetch_sources.py                      # 用默认路径，真下载
  python fetch_sources.py --dry-run            # 只预演：打印会做什么，不下载不写表
  python fetch_sources.py --xlsx path.xlsx     # 指定别的汇总表

做什么：
  1) 逐行读汇总表（列：id, topic, doc, org, year, url）
  2) 按 topic 挂领域前缀，生成规范文件名  <前缀>-<清洗后的文档名>.pdf
  3) 直链 PDF（url 以 .pdf 结尾，或响应头/魔数是 PDF）→ 下载到 raw-codex 隔离目录
     网页链接（text/html）→ 不下载，列进“待手动下载”清单，但表格行照写
  4) 每行写进一个【独立的待核验文件】registry_new.csv（格式与 registry.csv 完全一致），
     绝不直接改 registry.csv —— 由你人工核验后自行粘贴进去。
  5) 只读 registry.csv 用于「续接 id」和「按 url 去重」；已在库/已处理的行自动跳过，
     所以你以后往汇总表加新行、再跑一次，只会得到新增内容。

不做什么：不碰 registry.csv、不碰 evidence/raw/、不拆条、不判断 pages。
raw-codex 是未核验隔离区，你核验无误后再自行把文件搬到 evidence/raw/、把行粘进 registry.csv。
"""
import argparse
import csv
import os
import re
import sys
import time

import openpyxl
import requests

# Windows 控制台默认 GBK，emoji 和部分中文会 UnicodeEncodeError；强制 UTF-8 输出。
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))

# topic（汇总表里的领域名）→ 文件名前缀。未在此表中的 topic 会被跳过并报错，绝不瞎猜。
TOPIC_PREFIX = {
    "膳食补充剂与保健品": "ss",
    "减肥与代谢": "jf",
    "婴幼儿喂养儿童用药疫苗犹豫": "yyr",
    "癌症预防/防癌谣言": "az",
    # 备用（本批 xlsx 未出现，保留以便以后复用同一脚本）：
    "慢病常识": "mb",
    "用药误区": "yy",
    "睡眠": "sm",
    "孕产与产后护理": "yc",
    "孕期": "yq",
}

BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)

# Windows 非法文件名字符 → 全角替代（保持可读），其余控制字符直接删。
ILLEGAL_MAP = {
    "\\": "＼", "/": "／", ":": "：", "*": "＊",
    "?": "？", '"': "＂", "<": "＜", ">": "＞", "|": "｜",
}


def sanitize_filename(name: str, max_len: int = 120) -> str:
    """把文档名清洗成合法文件名主体（不含扩展名）。仅处理非法字符，中文原样保留。"""
    name = (name or "").strip()
    for bad, good in ILLEGAL_MAP.items():
        name = name.replace(bad, good)
    name = re.sub(r"[\x00-\x1f]", "", name)          # 去控制字符
    name = re.sub(r"\s+", " ", name).strip()          # 合并空白
    name = name.rstrip(". ")                            # Windows 不允许结尾的点/空格
    if len(name) > max_len:
        name = name[:max_len].rstrip(". ")
    return name or "未命名"


def load_registry_state(registry_path: str):
    """读现有 registry：返回 (最大id, 已有url集合, 已有文件名集合)。"""
    max_id = 0
    urls = set()
    filenames = set()
    if not os.path.exists(registry_path):
        return max_id, urls, filenames
    with open(registry_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                max_id = max(max_id, int((row.get("id") or "0").strip()))
            except ValueError:
                pass
            if row.get("url"):
                urls.add(row["url"].strip())
            if row.get("filename"):
                filenames.add(row["filename"].strip())
    return max_id, urls, filenames


def unique_filename(base_stem: str, ext: str, used: set) -> str:
    """在已用文件名集合里保证唯一：撞名就加 -2 / -3 …"""
    candidate = f"{base_stem}{ext}"
    n = 2
    while candidate in used:
        candidate = f"{base_stem}-{n}{ext}"
        n += 1
    used.add(candidate)
    return candidate


def looks_like_pdf_url(url: str) -> bool:
    return url.lower().split("?")[0].rstrip("/").endswith(".pdf")


def try_download_pdf(url: str, dest_path: str, timeout: int):
    """尝试下载。返回 ('pdf'|'html'|'fail', 说明)。仅当确认是 PDF 才落盘。"""
    try:
        with requests.get(
            url, headers={"User-Agent": BROWSER_UA}, timeout=timeout,
            stream=True, allow_redirects=True,
        ) as resp:
            if resp.status_code != 200:
                return "fail", f"HTTP {resp.status_code}"
            ctype = (resp.headers.get("Content-Type") or "").lower()
            first = next(resp.iter_content(chunk_size=1024), b"") or b""
            is_pdf = ("pdf" in ctype) or first[:5].startswith(b"%PDF")
            if not is_pdf:
                # 是网页（或其它），交给人工
                return "html", f"Content-Type={ctype or '未知'}"
            os.makedirs(os.path.dirname(dest_path), exist_ok=True)
            with open(dest_path, "wb") as out:
                out.write(first)
                for chunk in resp.iter_content(chunk_size=65536):
                    if chunk:
                        out.write(chunk)
            size = os.path.getsize(dest_path)
            if size < 1024:  # 太小八成是错误页伪装
                os.remove(dest_path)
                return "fail", f"文件过小({size}B)，疑似错误页"
            return "pdf", f"{round(size/1024/1024, 2)}MB"
    except Exception as e:
        return "fail", str(e)[:160]


def main():
    ap = argparse.ArgumentParser(description="deepresearch 汇总表 → 下载直链PDF + 填 registry.csv")
    ap.add_argument("--xlsx", default=os.path.join(os.path.dirname(BASE), "deepresearch", "数据.xlsx"))
    ap.add_argument("--out-dir", default=os.path.join(BASE, "evidence", "raw", "raw-codex"))
    ap.add_argument("--registry", default=os.path.join(BASE, "evidence", "registry.csv"),
                    help="只读：用于续接 id 和按 url 去重，绝不写入")
    ap.add_argument("--out", default=os.path.join(os.path.dirname(BASE), "deepresearch", "registry_new.csv"),
                    help="待核验输出文件（格式同 registry.csv），你核验后自行粘贴进 registry.csv")
    ap.add_argument("--timeout", type=int, default=30)
    ap.add_argument("--dry-run", action="store_true", help="只预演，不下载、不写表")
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        sys.exit(f"找不到汇总表：{args.xlsx}")

    wb = openpyxl.load_workbook(args.xlsx, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: header.index(name) for name in ("topic", "doc", "org", "year", "url") if name in header}
    missing = [c for c in ("topic", "doc", "org", "year", "url") if c not in col]
    if missing:
        sys.exit(f"汇总表缺列：{missing}，现有列={header}")

    max_id, existing_urls, existing_files = load_registry_state(args.registry)
    used_files = set(existing_files)
    next_id = max_id + 1

    new_rows = []          # 待写 registry 的行
    manual = []            # 待手动下载（网页）
    failed = []            # 下载失败
    skipped_existing = []  # url 已在库
    unknown_topic = []     # 未知领域

    for r in ws.iter_rows(min_row=2, values_only=True):
        topic = (str(r[col["topic"]]).strip() if r[col["topic"]] is not None else "")
        doc = (str(r[col["doc"]]).strip() if r[col["doc"]] is not None else "")
        org = (str(r[col["org"]]).strip() if r[col["org"]] is not None else "")
        year = (str(r[col["year"]]).strip() if r[col["year"]] is not None else "")
        url = (str(r[col["url"]]).strip() if r[col["url"]] is not None else "")
        if not (topic and doc and url):
            continue  # 空行

        if url in existing_urls:
            skipped_existing.append((doc, url))
            continue

        prefix = TOPIC_PREFIX.get(topic)
        if not prefix:
            unknown_topic.append((topic, doc))
            continue

        stem = f"{prefix}-{sanitize_filename(doc)}"
        filename = unique_filename(stem, ".pdf", used_files)
        dest = os.path.join(args.out_dir, filename)

        if args.dry_run:
            kind = "pdf?" if looks_like_pdf_url(url) else "网页?"
            print(f"[dry] {kind:5} {filename}  <=  {url}")
            new_rows.append([next_id, filename, org, doc, year, url, topic, ""])
            next_id += 1
            continue

        status, info = try_download_pdf(url, dest, args.timeout)
        if status == "pdf":
            print(f"[✅下载] {filename}  ({info})")
        elif status == "html":
            manual.append((filename, url, info))
            print(f"[📥手动] {filename}  <=  {url}  ({info})")
        else:
            failed.append((filename, url, info))
            print(f"[❌失败] {filename}  <=  {url}  ({info})")

        # 不管下载成功与否，registry 行照写（网页/失败的由你手动补文件到同名）
        new_rows.append([next_id, filename, org, doc, year, url, topic, ""])
        next_id += 1
        time.sleep(0.3)  # 轻微限速，别把对方服务器打急

    # ---- 写独立的待核验文件（绝不碰 registry.csv）----
    if new_rows and not args.dry_run:
        os.makedirs(os.path.dirname(args.out), exist_ok=True)
        # Excel on Windows only auto-detects UTF-8 reliably when a BOM is present.
        with open(args.out, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["id", "filename", "org", "doc", "year", "url", "topic", "pages"])
            for row in new_rows:
                w.writerow(row)
        print(f"\n[写入] 待核验文件 {args.out} 共 {len(new_rows)} 行（核验后自行粘进 registry.csv）")

    # ---- 待手动下载清单 ----
    if manual and not args.dry_run:
        todo = os.path.join(os.path.dirname(args.xlsx), "manual_download_todo.txt")
        with open(todo, "w", encoding="utf-8") as f:
            f.write("以下是网页链接（非直链PDF），请手动打开网页、保存/打印为 PDF，\n")
            f.write(f"并按【左边的文件名】存到：{args.out_dir}\n\n")
            for filename, url, info in manual:
                f.write(f"{filename}\t<=\t{url}\t({info})\n")
        print(f"[清单] 待手动下载 {len(manual)} 个 -> {todo}")

    # ---- 汇总 ----
    print("\n" + "=" * 50)
    print(f"总计新行 : {len(new_rows)}")
    print(f"✅ 已下载 : {len(new_rows) - len(manual) - len(failed)}" if not args.dry_run else f"(dry-run 未下载)")
    print(f"📥 待手动 : {len(manual)}")
    print(f"❌ 失败   : {len(failed)}")
    print(f"⏭  已存在 : {len(skipped_existing)}（url 已在库，跳过）")
    if unknown_topic:
        print(f"⚠ 未知领域(未处理) : {len(unknown_topic)} —— 请在 TOPIC_PREFIX 补映射：")
        for t, d in unknown_topic[:10]:
            print(f"    topic={t!r}  doc={d}")
    if failed:
        print("失败明细（可稍后重跑，已入库的会自动跳过）：")
        for fn, url, info in failed:
            print(f"    {fn}  ({info})")


if __name__ == "__main__":
    main()
